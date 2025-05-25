#!/usr/bin/env python
"""
Shift Report Data Updater

This script reads shift data from the BWP.XLSX file and generates a JavaScript data file
that will be used by the shift report web application.

Usage:
    1. Place your BWP.XLSX file in the 'data' folder (created automatically if needed)
    2. Run this script: python scripts/update_data.py
    3. The script will generate updated data in src/data/shifts.js

Requirements:
    - Python 3.6+
    - openpyxl module (for Excel processing)
"""

import os
import json
import sys
from datetime import datetime, timedelta
import uuid
import random

# Define paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'src', 'data')
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'shifts.js')

def ensure_directories_exist():
    """Ensure the data and output directories exist"""
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def try_import_openpyxl():
    """Try to import openpyxl, display helpful message if it's not installed"""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("\nERROR: The 'openpyxl' module is not installed.")
        print("Please install it using pip:")
        print("\npip install openpyxl\n")
        print("Then run this script again.\n")
        sys.exit(1)

def find_excel_file():
    """Find the BWP.XLSX file in the data directory"""
    # First look for BWP.XLSX
    for filename in os.listdir(DATA_DIR):
        if filename.upper() == "BWP.XLSX":
            return os.path.join(DATA_DIR, filename)
            
    # If not found, look for any Excel file
    excel_files = [f for f in os.listdir(DATA_DIR) if f.endswith(('.xlsx', '.xls'))]
    if not excel_files:
        print(f"\nNo Excel files found in the '{DATA_DIR}' directory.")
        print("Please place your BWP.XLSX file in this directory and run again.\n")
        print(f"Creating a sample BWP.XLSX file in {DATA_DIR} for demonstration.\n")
        create_sample_excel_file()
        return os.path.join(DATA_DIR, "BWP.xlsx")
    return os.path.join(DATA_DIR, excel_files[0])

def create_sample_excel_file():
    """Create a sample Excel file for demonstration"""
    openpyxl = try_import_openpyxl()
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shift Data"
    
    # Add headers based on the BWP.XLSX format
    headers = [
        "Id", "Completion time", "Operating hours (Hr)", "Final product (Tonnes)", 
        "Maximum flow reached (T/Hr)", "Frequency of stops", "ORE phosphate solids flowrate (T)",
        "Start-up preparation time (Hr)", "Total Ester consumption(L)", "Total Amin consumption(L)",
        "Total Acid consumption(L)", "Total Floculant consumption(m3)", "RECEIVED PHOSPHATE (T)",
        "Comment", "SHIFT", "RESPONSIBLE"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add sample data for last 14 days (7 days, 2 shifts per day)
    today = datetime.now()
    row = 2
    
    for i in range(13, -1, -1):  # 14 days including today
        date = today - timedelta(days=i)
        date_str = date.strftime("%d/%m/%Y")
        
        # Generate realistic data
        operating_hours = round(random.uniform(3, 10), 2)
        final_product = round(random.uniform(700, 5200), 2)
        max_flow = round(random.uniform(600, 700), 2)
        stops = round(random.uniform(3, 15))
        ore_flowrate = round(random.uniform(1000, 5500), 2)
        startup_time = round(random.uniform(2, 5), 2) if random.random() > 0.2 else "_"
        
        # Consumption data
        ester = round(random.uniform(20, 950), 2)
        amin = round(random.uniform(15, 950), 2)
        acid = round(random.uniform(20, 950), 2)
        floculant = round(random.uniform(3, 70), 2)
        
        # Sample comments
        comments = [
            "Normal operation with minor adjustments to flow rate.",
            "Conveyor SL4 down for maintenance for 45min.",
            "Trip of the recycle pump 230A-SP03 (low flow).",
            "Coarse rejects area experienced downtime across 3 events.",
            "Classification startup delayed due to valve issues.",
            "Scrubber repeatedly stopped and restarted, limiting stable feed.",
            "Trip of the raw water pump (motor fault).",
            "Clogging of the dehydration circuit.",
            "Blockage of the chute on the CAD conveyor."
        ]
        
        # Shift 1
        ws.cell(row=row, column=1, value=str(uuid.uuid4())[:8])
        ws.cell(row=row, column=2, value=date_str)
        ws.cell(row=row, column=3, value=operating_hours)
        ws.cell(row=row, column=4, value=final_product)
        ws.cell(row=row, column=5, value=max_flow)
        ws.cell(row=row, column=6, value=stops)
        ws.cell(row=row, column=7, value=ore_flowrate)
        ws.cell(row=row, column=8, value=startup_time)
        ws.cell(row=row, column=9, value=ester)
        ws.cell(row=row, column=10, value=amin)
        ws.cell(row=row, column=11, value=acid)
        ws.cell(row=row, column=12, value=floculant)
        ws.cell(row=row, column=13, value=0)  # RECEIVED PHOSPHATE
        ws.cell(row=row, column=14, value=random.choice(comments))
        ws.cell(row=row, column=15, value=1)  # SHIFT
        ws.cell(row=row, column=16, value=random.choice(["AKIL", "KANANE", "SADIK"]))
        
        row += 1
        
        # Shift 2
        ws.cell(row=row, column=1, value=str(uuid.uuid4())[:8])
        ws.cell(row=row, column=2, value=date_str)
        ws.cell(row=row, column=3, value=operating_hours * 0.9)  # Slightly different
        ws.cell(row=row, column=4, value=final_product * 0.95)  # Slightly different
        ws.cell(row=row, column=5, value=max_flow * 0.98)
        ws.cell(row=row, column=6, value=round(stops * 1.2))
        ws.cell(row=row, column=7, value=ore_flowrate * 0.97)
        ws.cell(row=row, column=8, value=startup_time)
        ws.cell(row=row, column=9, value=ester * 0.9)
        ws.cell(row=row, column=10, value=amin * 0.9)
        ws.cell(row=row, column=11, value=acid * 0.9)
        ws.cell(row=row, column=12, value=floculant * 0.9)
        ws.cell(row=row, column=13, value=0)  # RECEIVED PHOSPHATE
        ws.cell(row=row, column=14, value=random.choice(comments))
        ws.cell(row=row, column=15, value=2)  # SHIFT
        ws.cell(row=row, column=16, value=random.choice(["AKIL", "KANANE", "SADIK"]))
        
        row += 1
    
    # Save the workbook
    sample_file = os.path.join(DATA_DIR, "BWP.xlsx")
    wb.save(sample_file)
    print(f"Created sample Excel file: {sample_file}")
    
    return sample_file

def parse_date(date_str):
    """Parse a date string in DD/MM/YYYY format"""
    try:
        day, month, year = date_str.split('/')
        return f"{year}-{month}-{day}"
    except ValueError:
        print(f"Warning: Invalid date format: {date_str}. Expected DD/MM/YYYY.")
        return date_str

def process_excel_file(file_path):
    """Process the Excel file and extract shift data"""
    openpyxl = try_import_openpyxl()
    
    try:
        print(f"Processing Excel file: {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        shifts = []
        
        # Find headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers[cell_value.strip().lower()] = col
        
        # Required columns from BWP.XLSX
        expected_columns = [
            "completion time", "operating hours (hr)", "final product (tonnes)", 
            "maximum flow reached (t/hr)", "frequency of stops", "shift"
        ]
        
        missing_columns = [col for col in expected_columns if col not in headers]
        if missing_columns:
            print(f"WARNING: Missing expected columns in Excel file: {', '.join(missing_columns)}")
            print("Some data might not be available in the report.")
        
        # Read data rows
        for row in range(2, ws.max_row + 1):
            # Skip rows without date
            date_cell = ws.cell(row=row, column=headers.get("completion time", 0)).value
            if not date_cell:
                continue
            
            # Convert DD/MM/YYYY to YYYY-MM-DD for JavaScript
            date_str = parse_date(str(date_cell))
                
            # Extract shift data
            shift_data = {
                "id": str(row),
                "date": date_str,
                "shiftNumber": int(ws.cell(row=row, column=headers.get("shift", 0)).value or 1),
                "startTime": "07:00" if int(ws.cell(row=row, column=headers.get("shift", 0)).value or 1) == 1 else "19:00",
                "endTime": "19:00" if int(ws.cell(row=row, column=headers.get("shift", 0)).value or 1) == 1 else "07:00",
                "finalProductTonnes": float(ws.cell(row=row, column=headers.get("final product (tonnes)", 0)).value or 0),
                "operatingHours": float(ws.cell(row=row, column=headers.get("operating hours (hr)", 0)).value or 0),
                "maxFlowRate": float(ws.cell(row=row, column=headers.get("maximum flow reached (t/hr)", 0)).value or 0),
                "stopsFrequency": int(ws.cell(row=row, column=headers.get("frequency of stops", 0)).value or 0),
                "efficiency": calculate_efficiency(
                    float(ws.cell(row=row, column=headers.get("operating hours (hr)", 0)).value or 0),
                    int(ws.cell(row=row, column=headers.get("frequency of stops", 0)).value or 0)
                ),
                "downtime": calculate_downtime(
                    float(ws.cell(row=row, column=headers.get("operating hours (hr)", 0)).value or 0)
                ),
                "qualityRate": calculate_quality_rate(
                    float(ws.cell(row=row, column=headers.get("final product (tonnes)", 0)).value or 0),
                    float(ws.cell(row=row, column=headers.get("maximum flow reached (t/hr)", 0)).value or 0)
                ),
            }
            
            # Add optional fields if they exist
            optional_fields = [
                ("oreFlowrate", "ore phosphate solids flowrate (t)"),
                ("startupTime", "start-up preparation time (hr)"),
                ("esterConsumption", "total ester consumption(l)"),
                ("aminConsumption", "total amin consumption(l)"),
                ("acidConsumption", "total acid consumption(l)"),
                ("floculantConsumption", "total floculant consumption(m3)"),
                ("receivedPhosphate", "received phosphate (t)"),
                ("notes", "comment"),
                ("responsible", "responsible")
            ]
            
            for js_field, excel_field in optional_fields:
                if excel_field in headers:
                    cell_value = ws.cell(row=row, column=headers[excel_field]).value
                    if cell_value is not None and cell_value != "_":
                        shift_data[js_field] = str(cell_value) if js_field == "notes" or js_field == "responsible" else float(cell_value)
            
            shifts.append(shift_data)
        
        return shifts
        
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        return []

def calculate_efficiency(operating_hours, stops_frequency):
    """Calculate efficiency percentage based on operating hours and stops frequency"""
    if operating_hours <= 0:
        return 0
    
    # Simple efficiency calculation: higher operating hours and lower stops = higher efficiency
    # Max operating hours per shift is 12
    max_hours = 12
    max_stops = 20  # Assuming this is a reasonable maximum
    
    # Normalize operating hours (0-100%)
    hours_factor = min(operating_hours / max_hours, 1) * 80  # 80% weight to operating hours
    
    # Normalize stops (lower is better)
    stops_factor = (1 - min(stops_frequency / max_stops, 1)) * 20  # 20% weight to low stops
    
    return round(hours_factor + stops_factor, 1)

def calculate_downtime(operating_hours):
    """Calculate downtime in minutes based on operating hours"""
    max_hours = 12
    downtime_hours = max(0, max_hours - operating_hours)
    return round(downtime_hours * 60)  # Convert to minutes

def calculate_quality_rate(final_product, max_flow):
    """Calculate quality rate based on production and flow rate"""
    if max_flow <= 0:
        return 90  # Default quality rate
    
    # This is a simplified quality calculation
    # In a real system, you would have actual quality metrics
    theoretical_max = max_flow * 12  # Max flow over 12 hours
    
    if theoretical_max <= 0:
        return 90
    
    # Calculate what percentage of the theoretical maximum was achieved
    achievement_rate = min(final_product / theoretical_max, 1)
    
    # Scale to 80-100% quality range (assuming even low production has reasonable quality)
    quality_rate = 80 + (achievement_rate * 20)
    
    return round(quality_rate, 1)

def generate_js_file(shifts):
    """Generate a JavaScript file with the shift data"""
    js_content = f"""// Auto-generated file from BWP.XLSX
// Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
// DO NOT EDIT MANUALLY

export const shiftsData = {json.dumps(shifts, indent=2)};
"""
    
    with open(OUTPUT_FILE, 'w') as f:
        f.write(js_content)
    
    print(f"Generated JavaScript data file: {OUTPUT_FILE}")

def main():
    print("\n=== Shift Report Data Updater ===\n")
    
    # Ensure directories exist
    ensure_directories_exist()
    
    # Find Excel file
    excel_file = find_excel_file()
    
    # Process Excel file
    shifts = process_excel_file(excel_file)
    
    if not shifts:
        print("No valid shift data found in the Excel file.")
        return
    
    # Generate JavaScript file
    generate_js_file(shifts)
    
    print("\nData update complete!")
    print("The shift report web application will now show the updated data.")
    print("\nTo update the data again in the future:")
    print(f"1. Update your BWP.XLSX file in the '{DATA_DIR}' directory")
    print("2. Run this script again: python scripts/update_data.py")

if __name__ == "__main__":
    main()